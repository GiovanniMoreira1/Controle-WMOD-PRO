import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from credentials_app import DB_PASSWORD, DB_USER
from PIL import Image, ImageTk
import psycopg
import pandas as pd
import sys
import os
from tkextrafont import Font

def resource_path(relative_path): # solução encontrada para erro do PyInstaller não encontrar a imagem no caminhop
    try:
        base_path = sys._MEIPASS
    
    except Exception:
        base_path = os.path.abspath(".")
    
    return os.path.join(base_path, relative_path)


def menu_equipamentos():
    janela = tk.Toplevel()
    janela.title("Equipamentos")
    janela.attributes('-fullscreen', True)
    janela.bind('<Escape>', lambda event: janela.destroy())
    janela.configure(bg="#f4f6f9")
    
    
    container = tk.Frame(janela, bg="#f4f6f9")
    container.place(relx=0.5, rely=0.5, anchor="center")

    image_path = resource_path(os.path.join("app/assets", "icon.png"))
    img = Image.open(image_path)
    img = img.resize((120, 120))
    logo = ImageTk.PhotoImage(img)
    logo_label = tk.Label(container, image=logo)
    logo_label.image = logo
    logo_label.pack(side="top", pady=20)

    titulo = tk.Label(
        container,
        text="Gerenciamento de Equipamentos AUTOMOTIVE",
        font=custom_font_extrabold,
        fg="#00468e",
        bg="#f4f6f9"
    )
    titulo.pack(pady=(0, 30))
    janela.focus_force()

    def criar_botao(texto, comando):
        return tk.Button(
            container,
            text=texto,
            width=30,
            height=2,
            font=custom_font,
            bg="#00468e",
            fg="white",
            activebackground="#003366",
            activeforeground="white",
            relief="flat",
            cursor="hand2",
            command=comando
        )

    create_btn = criar_botao(
        "Cadastrar Equipamento",
        lambda: inserir_equipamento_janela(janela)
    )
    create_btn.pack(pady=10)

    delete_btn = criar_botao(
        "Excluir Equipamento",
        lambda: excluir_equipamento_janela(janela)
    )
    delete_btn.pack(pady=10)


    listar_btn = criar_botao(
        "Listar Equipamentos",
        lambda: listar_equipamentos_janela(janela)
    )
    listar_btn.pack(pady=10)

    update_btn = criar_botao(
        "Atualizar Equipamento",
        lambda: atualizar_equipamento_janela(janela)
    )
    update_btn.pack(pady=10)
    
    tabela_excel_btn = criar_botao(
        "Gerar Arquivo Excel",
        lambda: gerar_xlsx()
    )
    tabela_excel_btn.pack(pady=10)

    
def inserir_equipamento(params):
    with psycopg.connect(f"dbname=postgres user={DB_USER} password={DB_PASSWORD}") as conn:
        with conn.cursor() as cur:
            try:
                print("1")
                cur.execute(
                    "SELECT id_equipamento FROM equipamentos WHERE n_serie = %s", (params[4],)
                )
                print("2")
                id_equip = cur.fetchone()[0]
                if(id_equip != None):
                    print(id_equip)
                    print(params[1])
                    cur.execute(
                        "UPDATE equipamentos SET ativo = true, quantidade = quantidade + %s WHERE id_equipamento = %s", (params[1], id_equip,)
                    )
                    print("3")
                    
                    
                cur.execute(
                    "INSERT INTO equipamentos (nome, quantidade, descricao, fabricante, n_serie, localizacao, status, categoria) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)", (params)   
                )
                print("4")
                cur.execute(
                    f"SELECT id_equipamento FROM equipamentos WHERE nome = %s AND n_serie = %s", (params[0], params[4])
                )
                print("5")
                conn.commit()
            except Exception as e:
                print(e)
                return e
        
def inserir_equipamento_tk(params, janela, janela_pai):
    produto_criado = True
    try:
        e = inserir_equipamento(params)
        if(e) != None:
            produto_criado = False
            messagebox.showerror("Erro", f"Ocorreu um erro ao Cadastrar Produto: {e}. Tente novamente!")
        janela.destroy()
    finally:
        if(produto_criado):
            messagebox.showinfo("Sucesso!", "Produto cadastrado com sucesso!")
        janela_pai.deiconify()
        janela.destroy() 
       
def inserir_equipamento_janela(janela_pai):
    janela_pai.withdraw()

    janela = tk.Toplevel()
    janela.title("Inserir Equipamento")
    janela.geometry("1000x780")
    janela.configure(bg="#f4f6f8")
    janela.attributes('-fullscreen', True)
    janela.bind('<Escape>', lambda event: janela.destroy())
    janela.focus_set()

    name_var = tk.StringVar()
    qntd_var = tk.StringVar()
    fabr_var = tk.StringVar()
    loc_var = tk.StringVar()

    main_frame = tk.Frame(janela, bg="#f4f6f8")
    main_frame.pack(expand=True, fill="both")

    header = tk.Frame(main_frame, bg="#00468e", height=80)
    header.pack(fill="x")

    image_path = resource_path(os.path.join("app/assets", "icon.png"))
    img = Image.open(image_path)
    img = img.resize((120, 120))
    logo = ImageTk.PhotoImage(img)
    logo_label = tk.Label(header, image=logo, bg="#FFFFFF")
    logo_label.image = logo
    logo_label.pack(side="left")

    tk.Label(
        header,
        text="Cadastro de Equipamento",
        bg="#00468e",
        fg="white",
        font= custom_font_extrabold
    ).pack(pady=22)

    form_card = tk.Frame(
        main_frame,
        bg="white",
        highlightbackground="#d0d7de",
        highlightthickness=1
    )
    form_card.pack(pady=30, padx=50)

    label_style = {
        "bg": "white",
        "font": ("Segoe UI", 12, "bold"),
        "anchor": "e"
    }

    entry_style = {
        "font": ("Segoe UI", 11),
        "width": 35,
        "relief": "solid",
        "bd": 1
    }

    tk.Label(form_card, text="Nome:", **label_style).grid(row=0, column=0, padx=20, pady=12, sticky="e")
    name_entry = tk.Entry(form_card, textvariable=name_var, **entry_style)
    name_entry.grid(row=0, column=1, pady=12)

    tk.Label(form_card, text="Quantidade:", **label_style).grid(row=1, column=0, padx=20, pady=12, sticky="e")
    qntd_entry = tk.Entry(form_card, textvariable=qntd_var, **entry_style)
    qntd_entry.grid(row=1, column=1, pady=12)

    tk.Label(form_card, text="Fabricante:", **label_style).grid(row=2, column=0, padx=20, pady=12, sticky="e")
    fabr_entry = tk.Entry(form_card, textvariable=fabr_var, **entry_style)
    fabr_entry.grid(row=2, column=1, pady=12)

    tk.Label(form_card, text="Nº Série/Nº Modelo:", **label_style).grid(row=3, column=0, padx=20, pady=12, sticky="e")
    combo_nserie = ttk.Combobox(
        form_card,
        values=["Não se aplica"],
        width=35
    )
    combo_nserie.grid(row=3, column=1, pady=12)

    tk.Label(form_card, text="Localização:", **label_style).grid(row=4, column=0, padx=20, pady=12, sticky="e")
    loc_entry = tk.Entry(form_card, textvariable=loc_var, **entry_style)
    loc_entry.grid(row=4, column=1, pady=12)

    tk.Label(form_card, text="Status:", **label_style).grid(row=5, column=0, padx=20, pady=12, sticky="e")
    combo_status = ttk.Combobox(
        form_card,
        state="readonly",
        values=["Disponível", "Emprestado", "Com falha", "Não testado"],
        width=35
    )
    combo_status.grid(row=5, column=1, pady=12)

    tk.Label(form_card, text="Categoria:", **label_style).grid(row=6, column=0, padx=20, pady=12, sticky="e")
    combo_categ = ttk.Combobox(
        form_card,
        state="readonly",
        values=[
            "Equipamentos Principais",
            "Equipamentos de Rede",
            "Dispositivos de Campo",
            "Componentes elétricos e de painel",
            "Cabos",
            "Manuais e Guias",
            "Equipamentos Auxiliares",
            "Outros"
        ],
        width=35
    )
    combo_categ.grid(row=6, column=1, pady=12)

    tk.Label(form_card, text="Descrição:", **label_style).grid(
        row=7, column=0, padx=20, pady=12, sticky="ne"
    )

    desc_text = tk.Text(
        form_card,
        width=35,
        height=7,
        font=custom_font,
        relief="solid",
        bd=1,
        wrap="word"
    )
    desc_text.grid(row=7, column=1, pady=12)

    sub_btn = tk.Button(
        main_frame,
        text="Cadastrar Equipamento",
        font=custom_font_bold,
        bg="#00468e",
        fg="white",
        width=36,
        height=2,
        bd=0,
        cursor="hand2",
        activebackground="#003366",
        command=lambda: inserir_equipamento_tk(
            (
                name_entry.get().capitalize(),
                int(qntd_entry.get()),
                desc_text.get("1.0", "end").strip().capitalize(),
                fabr_entry.get().capitalize(),
                combo_nserie.get(),
                loc_entry.get().capitalize(),
                combo_status.get(),
                combo_categ.get()
            ),
            janela,
            janela_pai
        )
    )
    sub_btn.pack(pady=10)

def listar_equipamentos():
    with psycopg.connect(f"dbname=postgres user={DB_USER} password={DB_PASSWORD}") as conn:
        with conn.cursor() as cur:
            try:
                cur.execute(
                    "SELECT * FROM equipamentos WHERE ativo = true",   
                )
                conn.commit()

                return cur.fetchall()
            
            except Exception as e:
                return e
            
def excluir_equipamento_tk(tabela):
    selecionados = tabela.selection()
    iid = selecionados[0]

    if not selecionados:
        return
    try:
        excluir_equipamento(iid)
    except Exception as e:
        print(e)
    finally:
        if(tabela.exists(iid)):
            tabela.delete(iid)

def excluir_equipamento(id_equipamento):
    with psycopg.connect(f"dbname=postgres user={DB_USER} password={DB_PASSWORD}") as conn:
        with conn.cursor() as cur:
            try:
                cur.execute(
                    f"SELECT quantidade FROM equipamentos WHERE id_equipamento = %s", (id_equipamento,)
                )
                qntd = cur.fetchone()
                gerar_log("Retirada", qntd, id_equipamento)

                cur.execute(
                    f"UPDATE equipamentos SET ativo = FALSE WHERE id_equipamento = %s", (id_equipamento,),
                )
                conn.commit()
            except Exception as e:
                print(e)
                return e    
                
def excluir_equipamento_janela(janela_pai):
    janela_pai.withdraw()

    janela = tk.Toplevel(janela_pai)
    janela.attributes('-fullscreen', True)
    janela.bind('<Escape>', lambda event: janela.destroy())
    janela.title("Excluir Equipamento")
    janela.geometry("1000x650")
    janela.configure(bg="#f4f6f8")
    janela.focus_set()

    main_frame = tk.Frame(janela, bg="#f4f6f8")
    main_frame.pack(expand=True, fill="both")

    header = tk.Frame(main_frame, bg="#00468e", height=80)
    header.pack(fill="x")
    header.pack_propagate(False)

    tk.Label(
        header,
        text="Excluir Equipamento",
        bg="#00468e",
        fg="white",
        font=custom_font_extrabold
    ).pack(side="left", padx=30)

    card = tk.Frame(
        main_frame,
        bg="white",
        highlightbackground="#d0d7de",
        highlightthickness=1
    )
    card.pack(padx=40, pady=30, fill="both", expand=True)
    content = tk.Frame(card, bg="white")
    content.pack(padx=20, pady=20, fill="both", expand=True)

    tk.Label(
        content,
        text="Selecione um equipamento para excluir",
        bg="white",
        font=custom_font_bold,
    ).pack(anchor="w", pady=(0, 10))

    table_frame = tk.Frame(content, bg="white")
    table_frame.pack(fill="both", expand=True)

    scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL)

    colunas = ("nome", "quantidade", "fabricante", "n_serie", "localizacao", "status")

    tabela = ttk.Treeview(
        table_frame,
        columns=colunas,
        show="headings",
        yscrollcommand=scrollbar.set,
        selectmode="browse",
        height=12
    )

    scrollbar.config(command=tabela.yview)

    tabela.heading("nome", text="Nome")
    tabela.heading("quantidade", text="Qtd")
    tabela.heading("fabricante", text="Fabricante")
    tabela.heading("n_serie", text="Nº Série/Nº Modelo")
    tabela.heading("localizacao", text="Localização")
    tabela.heading("status", text="Status")

    tabela.column("nome", width=220)
    tabela.column("quantidade", width=60, anchor="center")
    tabela.column("fabricante", width=160)
    tabela.column("n_serie", width=130)
    tabela.column("localizacao", width=120, anchor="center")
    tabela.column("status", width=120, anchor="center")

    rows = listar_equipamentos()
    for row in rows:
        tabela.insert(
            "",
            tk.END,
            iid=row[0],
            values=(
                row[1],
                row[2],
                row[4],
                row[5],
                row[6],
                row[7]
            )
        )

    tabela.pack(side=tk.LEFT, fill="both", expand=True)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    actions = tk.Frame(content, bg="white")
    actions.pack(fill="x", pady=15)

    delete_btn = tk.Button(
        actions,
        text="🗑  Excluir Equipamento",
        font=custom_font_bold,
        bg="#c0392b",
        fg="white",
        width=28,
        height=2,
        bd=0,
        cursor="hand2",
        activebackground="#922b21",
        activeforeground="white",
        command=lambda: excluir_equipamento_tk(tabela)
    )
    delete_btn.pack(side="right")

def retornar_tela(janela_pai, janela):
    janela.withdraw()
    janela_pai.focus_set()

def listar_equipamentos_janela(janela_pai):
    janela_pai.withdraw()

    janela = tk.Toplevel(janela_pai)
    janela.attributes('-fullscreen', True)
    janela.bind('<Escape>', lambda event: janela.destroy())
    janela.title("Listar Equipamentos")
    janela.geometry("1000x650")
    janela.configure(bg="#f4f6f8")
    janela.focus_set()

    main_frame = tk.Frame(janela, bg="#f4f6f8")
    main_frame.pack(expand=True, fill="both")

    header = tk.Frame(main_frame, bg="#00468e", height=80)
    header.pack(fill="x")
    header.pack_propagate(False)

    tk.Label(
        header,
        text="Listar Equipamentos",
        bg="#00468e",
        fg="white",
        font=custom_font_extrabold,
    ).pack(side="left", padx=30)

    card = tk.Frame(
        main_frame,
        bg="white",
        highlightbackground="#d0d7de",
        highlightthickness=1
    )
    card.pack(padx=40, pady=30, fill="both", expand=True)
    content = tk.Frame(card, bg="white")
    content.pack(padx=20, pady=20, fill="both", expand=True)

    table_frame = tk.Frame(content, bg="white")
    table_frame.pack(fill="both", expand=True)

    scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL)

    colunas = ("nome", "quantidade", "fabricante", "n_serie", "localizacao", "status")

    tabela = ttk.Treeview(
        table_frame,
        columns=colunas,
        show="headings",
        yscrollcommand=scrollbar.set,
        selectmode="browse",
        height=12
    )

    scrollbar.config(command=tabela.yview)

    tabela.heading("nome", text="Nome")
    tabela.heading("quantidade", text="Qtd")
    tabela.heading("fabricante", text="Fabricante")
    tabela.heading("n_serie", text="Nº Série/Nº Modelo")
    tabela.heading("localizacao", text="Localização")
    tabela.heading("status", text="Status")

    tabela.column("nome", width=220)
    tabela.column("quantidade", width=60, anchor="center")
    tabela.column("fabricante", width=160)
    tabela.column("n_serie", width=130)
    tabela.column("localizacao", width=120, anchor="center")
    tabela.column("status", width=120, anchor="center")

    rows = listar_equipamentos()
    for row in rows:
        tabela.insert(
            "",
            tk.END,
            iid=row[0],
            values=(
                row[1],
                row[2],
                row[4],
                row[5],
                row[6],
                row[7]
            )
        )

    tabela.pack(side=tk.LEFT, fill="both", expand=True)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    actions = tk.Frame(content, bg="white")
    actions.pack(fill="x", pady=15)

    voltar_btn = tk.Button(
        actions,
        text="Voltar",
        font=custom_font_bold,
        bg="#c0392b",
        fg="white",
        width=28,
        height=2,
        bd=0,
        cursor="hand2",
        activebackground="#922b21",
        activeforeground="white",
        command=lambda: retornar_tela(janela_pai, janela)
    )
    voltar_btn.pack(side="right")

def chamar_erro(erro):
    tk.messagebox.showerror(
        "Erro",
        f"Não foi possível concluir a operação. Erro: {erro}"
    )

def refresh_treeview(tabela): #func p dar refresh na tabela apos alterações
    tabela.delete(*tabela.get_children())
    rows = listar_equipamentos()
    for row in rows:
        tabela.insert(
            "",
            "end",
            iid=row[0],
            values=(
                row[1],
                row[2],
                row[4],
                row[5],
                row[6],
                row[7]
            )
        )

def atualizar_equipamento(iid, categoria, janela, tabela):
     novo_nome = simpledialog.askstring(
         f"Atualizar {categoria}",
         f"Digite o novo valor para {categoria}",
         parent=janela
     )
     with psycopg.connect(f"dbname=postgres user={DB_USER} password={DB_PASSWORD}") as conn:
        with conn.cursor() as cur:
            try:
                if(categoria == "quantidade"):
                    novo_nome = int(novo_nome)
                    
                cur.execute(
                    f"UPDATE equipamentos SET {categoria} = %s WHERE id_equipamento = %s", (novo_nome, iid)
                )
                cur.execute(
                    f"SELECT quantidade FROM equipamentos WHERE id_equipamento = %s", (iid,)
                )
                quantidade = cur.fetchone()
                conn.commit()

                tk.messagebox.showinfo(
                    "Informação",
                    "Equipamento atualizado com sucesso!",
                    parent=janela
                )
                gerar_log("Alteração", quantidade, iid)
                janela.destroy()
                atualizar_equipamento_tk(tabela)
            except Exception as e:
                chamar_erro(e)
                return e
            
def buscar_nome(iid, categoria):
     with psycopg.connect(f"dbname=postgres user={DB_USER} password={DB_PASSWORD}") as conn:
        with conn.cursor() as cur:
            try:
                cur.execute(
                    f"SELECT {categoria} FROM equipamentos WHERE id_equipamento = %s", (iid,)
                )
                return cur.fetchone()
            except Exception as e:
                print(e)
                return e 

def atualizar_equipamento_tk(tabela):
    janela = tk.Toplevel(tabela)
    janela.attributes('-fullscreen', True)
    janela.bind('<Escape>', lambda event: janela.destroy())
    janela.title("Atualizar Equipamento")
    janela.geometry("1000x650")
    janela.configure(bg="#f1f4f9")
    janela.focus_set()

    selecionados = tabela.selection()
    if selecionados != ():
        iid = selecionados[0]

        container = tk.Frame(
            janela,
            bg="#ffffff",
            bd=1,
            relief="solid",
            highlightbackground="#e5e7eb",
            highlightthickness=1
        )
        container.pack(padx=50, pady=40, fill="both", expand=True)

        titulo = tk.Label(
            container,
            text="Atualizar Dados do Equipamento",
            font=custom_font_extrabold,
            bg="#ffffff",
            fg="#1f2937"
        )
        titulo.grid(row=0, column=0, columnspan=3, pady=(25, 35))

        def linha(label_texto, valor, linha, categoria):
            tk.Label(
                container,
                text=f"{label_texto}:",
                font=custom_font,
                bg="#ffffff",
                fg="#374151",
                width=15,
                anchor="e"
            ).grid(row=linha, column=0, padx=(30, 12), pady=10)

            tk.Label(
                container,
                text=valor,
                font=custom_font,
                bg="#ffffff",
                fg="#111827",
                anchor="w",
                wraplength=420,
                justify="left"
            ).grid(row=linha, column=1, sticky="w", pady=10)

            btn = tk.Button(
                container,
                text="Atualizar",
                width=18,
                bg="#00468e",
                fg="white",
                activebackground="#005fbf",
                activeforeground="white",
                relief="flat",
                cursor="hand2",
                command=lambda: atualizar_equipamento(iid, categoria, janela, tabela)
            )
            btn.grid(row=linha, column=2, padx=30, pady=10)

        linha("Nome", buscar_nome(iid, "nome"), 1, "nome")
        linha("Quantidade", buscar_nome(iid, "quantidade"), 2, "quantidade")
        linha("Descrição", buscar_nome(iid, "descricao"), 3, "descricao")
        linha("Fabricante", buscar_nome(iid, "fabricante"), 4, "fabricante")
        linha("Nº Série/Nº Modelo", buscar_nome(iid, "n_serie"), 5, "n_serie")
        linha("Localização", buscar_nome(iid, "localizacao"), 6, "localizacao")

        container.grid_rowconfigure(8, weight=1)
    else:
        tk.messagebox.showwarning("Warning", "Selecione um produto para atualizar!", parent=janela)
        janela.destroy()
        return
    
def atualizar_equipamento_janela(janela_pai):
    janela_pai.withdraw()

    janela = tk.Toplevel(janela_pai)
    janela.attributes('-fullscreen', True)
    janela.bind('<Escape>', lambda event: janela.destroy())
    janela.title("Atualizar Equipamento")
    janela.geometry("1000x650")
    janela.configure(bg="#f4f6f8")
    janela.focus_set()

    main_frame = tk.Frame(janela, bg="#f4f6f8")
    main_frame.pack(expand=True, fill="both")

    header = tk.Frame(main_frame, bg="#00468e", height=80)
    header.pack(fill="x")
    header.pack_propagate(False)

    tk.Label(
        header,
        text="Atualizar Equipamento",
        bg="#00468e",
        fg="white",
        font=custom_font_extrabold,
    ).pack(side="left", padx=30)

    card = tk.Frame(
        main_frame,
        bg="white",
        highlightbackground="#d0d7de",
        highlightthickness=1
    )
    card.pack(padx=40, pady=30, fill="both", expand=True)
    content = tk.Frame(card, bg="white")
    content.pack(padx=20, pady=20, fill="both", expand=True)

    tk.Label(
        content,
        text="Selecione um equipamento para atualizar",
        bg="white",
        font=custom_font_bold,
    ).pack(anchor="w", pady=(0, 10))

    table_frame = tk.Frame(content, bg="white")
    table_frame.pack(fill="both", expand=True)

    scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL)

    colunas = ("nome", "quantidade", "fabricante", "n_serie", "localizacao", "status")

    tabela = ttk.Treeview(
        table_frame,
        columns=colunas,
        show="headings",
        yscrollcommand=scrollbar.set,
        selectmode="browse",
        height=12
    )

    scrollbar.config(command=tabela.yview)

    tabela.heading("nome", text="Nome")
    tabela.heading("quantidade", text="Qntd")
    tabela.heading("fabricante", text="Fabricante")
    tabela.heading("n_serie", text="Nº Série/Nº Modelo")
    tabela.heading("localizacao", text="Localização")
    tabela.heading("status", text="Status")

    tabela.column("nome", width=220)
    tabela.column("quantidade", width=60, anchor="center")
    tabela.column("fabricante", width=160)
    tabela.column("n_serie", width=130)
    tabela.column("localizacao", width=120, anchor="center")
    tabela.column("status", width=120, anchor="center")

    rows = listar_equipamentos()
    for row in rows:
        tabela.insert(
            "",
            tk.END,
            iid=row[0],
            values=(
                row[1],
                row[2],
                row[4],
                row[5],
                row[6],
                row[7]
            )
        )

    tabela.pack(side=tk.LEFT, fill="both", expand=True)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    actions = tk.Frame(content, bg="white")
    actions.pack(fill="x", pady=15)
    update_btn = tk.Button(
        actions,
        text="🔃  Atualizar Equipamento",
        font=custom_font_bold,
        bg="#212c92",
        fg="white",
        width=28,
        height=2,
        bd=0,
        cursor="hand2",
        activebackground="#212c92",
        activeforeground="white",
        command=lambda: atualizar_equipamento_tk(tabela)
    )
    refresh_btn = tk.Button(
        actions,
        text="Refresh",
        font=custom_font_bold,
        bg="#212c92",
        fg="white",
        width=28,
        height=2,
        bd=0,
        cursor="hand2",
        activebackground="#212c92",
        activeforeground="white",
        command=lambda: {
            atualizar_equipamento_janela(janela_pai),
            janela.destroy()
        }
    )
    update_btn.pack(side="right")
    refresh_btn.pack(side="right", padx=15)

def gerar_log(operacao, quantidade, id_equip):
    with psycopg.connect(f"dbname=postgres user={DB_USER} password={DB_PASSWORD}") as conn:
        with conn.cursor() as cur:
            try:
                cur.execute(
                    "INSERT INTO operacoes (tipo_operacao, id_funcionario, id_equipamento, quantidade_op) VALUES (%s, %s, %s, %s)", (operacao, 1, id_equip, int(quantidade[0]))
                )
                conn.commit()
            except Exception as e:
                print(e)
                return e
    
def gerar_xlsx():
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
    
    sql_query = "SELECT * FROM equipamentos WHERE ativo = true"
    
    with psycopg.connect(f"dbname=postgres user={DB_USER} password={DB_PASSWORD}") as conn:
        df = pd.read_sql_query(sql_query, con=conn)
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="ID")
        ws.cell(row=1, column=2, value="Nome")
        ws.cell(row=1, column=3, value="Quantidade")
        ws.cell(row=1, column=4, value="Descrição")
        ws.cell(row=1, column=5, value="Fabricante")
        ws.cell(row=1, column=6, value="Nº Série/Nº Modelo")
        ws.cell(row=1, column=7, value="Localização")
        ws.cell(row=1, column=8, value="Status")
        ws.cell(row=1, column=9, value="Categoria")
        ws.cell(row=1, column=10, value="Ativo")
        ws.column_dimensions['B'].width = 56
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 120
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 28
        ws.column_dimensions['G'].width = 13
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 34
        ws.column_dimensions['J'].width = 13
        ws.title = "Inventário WinMOD PRO"
        
        for row in df.values.tolist():
            ws.append(row)
            
        end_cell = get_column_letter(ws.max_column) + str(ws.max_row)
        table_ref = f"A1:{end_cell}"
        
        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        table = Table(displayName="WinMOD_PRO", ref=table_ref)
        table.tableStyleInfo = style
        
        ws.add_table(table)
        
        file_name = 'WinMOD PRO - Inventário.xlsx'
        wb.save(file_name) 
    
root = tk.Tk()
custom_font = Font(file="app/fonts/Barlow-Black.ttf", family="Barlow", size=12)
custom_font_bold = Font(file="app/fonts/Barlow-Bold.ttf", family="Barlow", size=15)
custom_font_extrabold = Font(file="app/fonts/Barlow-ExtraBold.ttf", family="Barlow", size=20)
root.attributes('-fullscreen', True)
root.bind('<Escape>', lambda event: root.destroy())
root.title("Sistema de Controle")
root.geometry("900x600")
root.configure(bg="#f4f6f8")

main_frame = tk.Frame(root, bg="#f4f6f8")
main_frame.pack(expand=True, fill="both")

header = tk.Frame(main_frame, bg="#00468e", height=90)
header.pack(fill="x")

header.pack_propagate(False)
image_path = resource_path(os.path.join("app/assets", "icon.png"))
img = Image.open(image_path)
img = img.resize((120, 120))
logo = ImageTk.PhotoImage(img)
logo_label = tk.Label(header, image=logo, bg="#FFFFFF")
logo_label.image = logo
logo_label.pack(side="right")

title = tk.Label(
    header,
    text="Controle WinMOD PRO",
    bg="#00468e",
    fg="white",
    font=custom_font_extrabold,
)
title.pack(side="left", padx=30)

card = tk.Frame(
    main_frame,
    bg="white",
    highlightbackground="#d0d7de",
    highlightthickness=1
)
card.pack(pady=80, padx=200)

card_content = tk.Frame(card, bg="white")
card_content.pack(padx=40, pady=40)

subtitle = tk.Label(
    card_content,
    text="Menu Principal",
    bg="white",
    font=custom_font_bold,
)
subtitle.pack(pady=(0, 30))

sub_btn = tk.Button(
    card_content,
    text="⚙️  Menu de Equipamentos",
    font=custom_font_bold,
    bg="#00468e",
    fg="white",
    width=28,
    height=2,
    bd=0,
    cursor="hand2",
    activebackground="#003366",
    activeforeground="white",
    command=menu_equipamentos
)
sub_btn.pack()

footer = tk.Label(
    main_frame,
    text="© Sistema Interno • Controle WinMOD PRO",
    bg="#f4f6f8",
    fg="#6b7280",
    font=custom_font
)
footer.pack(side="bottom", pady=10)

root.mainloop()
