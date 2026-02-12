from credentials import DB_PASSWORD, DB_USER, EMAIL, EMAIL_PASSWORD, EMAIL_ENVIO
import psycopg
import pandas as pd
import os
import time
import email
import json
import imaplib
from redmail import EmailSender
from imap_tools import MailBox
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


def alteracao_bd_json(): # função responsavel por realizar as operações no banco de dados
    
    with open('dados.json', 'r', encoding='utf-8-sig') as f: # dados.json é o arquivo gerado e enviado ao e-mail
        data = json.load(f)
        
    df = pd.DataFrame([data]) # cria um dataframe com os dados que vierem do email (forms -> email -> python)
    
    if(not esta_cadastrado(df['email_funcionario'][0])): # chama função com valor do e-mail presente no forms com intuito de checar se o funcionário já esta cadastrado no sistema, caso não esteja, ocorre o cadastro
        try: 
            with psycopg.connect(f"dbname=postgres user={DB_USER} password={DB_PASSWORD}") as conn:
                with conn.cursor() as cur:
                        cur.execute(
                            "INSERT INTO funcionarios (nome, email) VALUES (%s, %s)", (df['nome_funcionario'][0], df['email_funcionario'][0]) # insere o funcionario no banco de dados
                        )
                        conn.commit()
                        
        except Exception as e:
            print(f"Erro ao cadastrar user no banco de dados. Erro: {e}")
            return
    
    if(df['operacao'][0] == "Entrada"): # caso da operação ser entrada
        print(f"""
            =-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
              Operação: Entrada
              
              Nome: {df['nome_funcionario'][0]}
              E-mail: {df['email_funcionario'][0]}
              Equipamento: {df['nome_equipamento'][0]}
              Quantidade: {df['quantidade'][0]}
              Armário: {df['nome_armario'][0]}
            =-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
              """)
        if(not nome_ja_cadastrado(df["nome_equipamento"][0])): # chama função com nome do equipamento presente no forms com intuito de checar se o equipamento já esta cadastrado no sistema, caso não esteja, ocorre o cadastro.
            params = (df['nome_equipamento'][0], int(df['quantidade'][0]), '', '', '', df['nome_armario'][0], "Disponível", df['categoria'][0])
            try: 
                with psycopg.connect(f"dbname=postgres user={DB_USER} password={DB_PASSWORD}") as conn:
                    with conn.cursor() as cur:
                            cur.execute( # insere o equipamento na tabela de equipamentos
                                "INSERT INTO equipamentos (nome, quantidade, descricao, fabricante, n_serie, localizacao, status, categoria) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)", (params)   
                            )
            except Exception as e:
                enviar_email_erro(e, df['email_funcionario'][0])
                return
                    
        else: # caso o equipamento já esteja armazenado no sistema, é utilizado o valor da requisição para acrescentar a quantidade no banco de dados (valor_antigo + requisicao_atual)
            try: 
                with psycopg.connect(f"dbname=postgres user={DB_USER} password={DB_PASSWORD}") as conn:
                    with conn.cursor() as cur:
                            cur.execute(
                                "UPDATE equipamentos SET quantidade = quantidade + %s WHERE nome = %s", (int(df['quantidade'][0]), df['nome_equipamento'][0])
                            )
            except Exception as e:
                enviar_email_erro(e, df['email_funcionario'][0])
                return
            
    elif(df['operacao'][0] == "Devolução"): # caso da operação ser devolução
        print(f"""
            =-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
              Operação: Devolução
              
              Nome: {df['nome_funcionario'][0]}
              E-mail: {df['email_funcionario'][0]}
              Equipamento: {df['nome_equipamento'][0]}
              Quantidade: {df['quantidade'][0]}
              Armário: {df['nome_armario'][0]}
            =-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
              """)
        if(not nome_ja_cadastrado(df['nome_equipamento'][0])): # caso não exista um equipamento no banco de dados com o mesmo nome
            params = (df['nome_equipamento'][0], int(df['quantidade'][0]), '', '', '', df['nome_armario'][0], "Disponível", df['categoria'][0])
            try: 
                with psycopg.connect(f"dbname=postgres user={DB_USER} password={DB_PASSWORD}") as conn:
                    with conn.cursor() as cur: # insere o equipamento no bd
                            cur.execute(
                                "INSERT INTO equipamentos (nome, quantidade, descricao, fabricante, n_serie, localizacao, status, categoria) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)", (params)   
                            )
            except Exception as e:
                enviar_email_erro(e, df['email_funcionario'][0])
                return
                    
        else: # caso já exista um equipamento com o mesmo nome, a quantidade do banco de dados é alterada somando a quantidade dos itens que vão entrar, dessa maneira, evitando duplicidade nos dados
            try: 
                with psycopg.connect(f"dbname=postgres user={DB_USER} password={DB_PASSWORD}") as conn:
                    with conn.cursor() as cur:
                            cur.execute(
                                "UPDATE equipamentos SET quantidade = quantidade + %s, status = %s WHERE nome = %s", (int(df['quantidade'][0]), 'Disponível', df['nome_equipamento'][0])
                            )
                            cur.execute(
                                "UPDATE emprestimos SET status_atual = %s WHERE id_equipamento = %s", ('Devolvido', df['id_equipamento'][0]) 
                            )
            except Exception as e:
                enviar_email_erro(e, df['email_funcionario'][0])
                return
            
        
    elif(df['operacao'][0] == "Retirada"): # caso da operação ser retirada
        print(f"""
            =-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
              Operação: Retirada
              
              Nome: {df['nome_funcionario'][0]}
              E-mail: {df['email_funcionario'][0]}
              Equipamento: {df['nome_equipamento'][0]}
              Quantidade: {df['quantidade'][0]}
              Armário: {df['nome_armario'][0]}
            =-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
              """)
        try: 
            with psycopg.connect(f"dbname=postgres user={DB_USER} password={DB_PASSWORD}") as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT quantidade FROM equipamentos WHERE id_equipamento = %s", (df['id_equipamento'][0],)  # query para pegar a quantidade disponivel no bd
                    )
                    quantidade_bd = int((cur.fetchone()[0]))
                    
                    if quantidade_bd == int(df['quantidade'][0]): # caso 1: quantidade do equipamento no banco de dados ser exatamente a mesma da solicitada para retirada
                        cur.execute(
                            "UPDATE equipamentos SET ativo = false WHERE id_equipamento = %s", (df['id_equipamento'][0],) 
                        )
                    elif quantidade_bd < int(df['quantidade'][0]): # caso 2: quantidade do equipamento no banco de dados ser insuficiente se comparada com a solicitada para retirada
                        raise Exception
                    elif quantidade_bd > int(df['quantidade'][0]): # caso 3: quantidade do equipamento no banco de dados ser maior que a solicitada para retirada
                        cur.execute(
                            "UPDATE equipamentos SET quantidade = quantidade - %s WHERE id_equipamento = %s", (df['quantidade'][0], df['id_equipamento'][0])
                        )
        
                        
        except Exception as e:
            enviar_email_erro(e, df['email_funcionario'][0])
            return
            
    
    elif(df['operacao'][0] == "Empréstimo"):
        print(f"""
            =-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
              Operação: Empréstimo
              
              Nome: {df['nome_funcionario'][0]}
              E-mail: {df['email_funcionario'][0]}
              Equipamento: {df['nome_equipamento'][0]}
              Quantidade: {df['quantidade'][0]}
              Armário: {df['nome_armario'][0]}
            =-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
              """)
        try:
            with psycopg.connect(f"dbname=postgres user={DB_USER} password={DB_PASSWORD}") as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT status FROM equipamentos WHERE id_equipamento = %s", (df['id_equipamento'][0],)  # query para pegar o status do equipamento no bd
                    )
                    status = cur.fetchone()[0]
                    quantidade_forms = df['quantidade'][0]
                        # query para pegar o valor da quantidade presente no banco de dados e fazer a comparação com o pedido (if)
                    if(quantidade_forms == df['quantidade'][0]):
                        cur.execute(
                            "UPDATE equipamentos SET quantidade = 0, status = 'Emprestado' WHERE id_equipamento = %s", (df['id_equipamento'][0],)
                        )
                        cur.execute( # resgatar ID do funcionário
                            "SELECT id_funcionario FROM funcionarios WHERE email = %s", (df['email_funcionario'][0],)
                        )
                        id_func = cur.fetchone()[0]
                        parametros = (id_func, df['id_equipamento'][0], df['data_retorno'][0], quantidade_forms, 'Emprestado')
                        print(parametros)
                        cur.execute( 
                            "INSERT INTO emprestimos (id_funcionario, id_equipamento, data_retorno, quantidade, status_atual) VALUES (%s, %s, %s, %s, %s)", parametros 
                        )
                    elif(quantidade_forms < df['quantidade'][0]): # caso 2: quantidade do equipamento no banco de dados ser maior que a solicitada para emprestimo
                        cur.execute( # update na quantidade 
                            "UPDATE equipamentos SET quantidade = quantidade - %s WHERE id_equipamento = %s", (quantidade_forms, df['id_equipamento'][0])
                        )

                        cur.execute( # resgatar ID do funcionário
                            "SELECT id_funcionario FROM funcionarios WHERE email = %s", (df['email_funcionario'][0],)
                        )
                        id_func = cur.fetchone()[0]
                                                    
                        parametros = (id_func, df['id_equipamento'][0], df['data_retorno'][0], quantidade_forms, 'Emprestado')      
                        cur.execute( 
                            "INSERT INTO emprestimos (id_funcionario, id_equipamento, data_retorno, quantidade, status_atual) VALUES (%s, %s, %s, %s, %s)", parametros 
                        )
                            
                    else: # caso 3: quantidade do equipamento no banco de dados ser insuficiente se comparada com a solicitada para retirada (erro)
                        raise Exception
        except Exception as e:
            #enviar_email_erro(e, df['email_funcionario'][0])
            print(e)
            return
        gerar_xlsx_emprestimos()

    gerar_xlsx()
    gerar_xlsx_operacoes()
    enviar_email_sucesso(df['operacao'][0], df['nome_equipamento'][0], df['email_funcionario'][0])
    insert_operacao_bd(df['email_funcionario'][0], df['operacao'][0], df['quantidade'][0], df['id_equipamento'][0])    
    return
        
def esperar_email(): 
    with MailBox('imap.gmail.com').login(EMAIL, EMAIL_PASSWORD) as mailbox:
        print("Aguardando e-mail...")
        responses = mailbox.idle.wait(timeout=600) # entra em modo idle (aguarda qualquer e-mail novo)
        
        if responses:
            for resp in responses:
                if 'EXISTS' in str(resp): # recebe qualquer e-mail no inbox
                    print("Novo e-mail chegou!")
                    ler_anexo()


def ler_anexo():
    mail = imaplib.IMAP4_SSL('imap.gmail.com')
    mail.login(EMAIL, EMAIL_PASSWORD)
    mail.select('inbox')
    
    criterio_busca = f"UNSEEN FROM {EMAIL_ENVIO}"
    status, data = mail.search(None, criterio_busca)
    
    email_ids = data[0].split()
    
    for email_id in email_ids: # separa os emails em IDs
        status, msg_data = mail.fetch(email_id, '(RFC822)')
        for response_part in msg_data:
            if(isinstance(response_part, tuple)):
                msg = email.message_from_bytes(response_part[1])
            
            for part in msg.walk():
                if(part.get_content_maintype == "multipart"):
                    continue
                if(part.get_content_disposition is None):
                    continue
                
                filename = part.get_filename()
                if filename:
                    filepath = os.path.join(os.getcwd(), filename)
                    with open(filepath, "wb") as f:
                        f.write(part.get_payload(decode=True))
                        
    mail.close()
    mail.logout()
    
    alteracao_bd_json()

def esta_cadastrado(email):
    try:
        with psycopg.connect(f"dbname=postgres user={DB_USER} password={DB_PASSWORD}") as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id_funcionario FROM funcionarios WHERE email = %s", (email,) # busca o id que esta cadastrado com o email
                )    
                conn.commit()
                if(cur.fetchone() == None):
                    return False
                else:
                    return True
    except Exception as e:
        print("erro 4")
        print(email)
        print(e)
        return
            
def nome_ja_cadastrado(nome):
    try:
        with psycopg.connect(f"dbname=postgres user={DB_USER} password={DB_PASSWORD}") as conn:
            with conn.cursor() as cur:
                    cur.execute(
                        "SELECT id_equipamento FROM equipamentos WHERE nome = %s", (nome,) # busca o id que esta cadastrado com o email
                    )    
                    conn.commit()
                    
                    if(cur.fetchone() == None):
                        return False
                    else:
                        return True
                    
    except Exception as e:
        print("erro 3")
        print(nome)
        print(e)
        return

def insert_operacao_bd(email, operacao, quantidade, id_equip): # função responsavel por inserir as operações na sua respectiva tabela
    try:
        with psycopg.connect(f"dbname=postgres user={DB_USER} password={DB_PASSWORD}") as conn:
            with conn.cursor() as cur:
                if(operacao == "Entrada"):
                    operacao = "Entrada"
                    cur.execute( # pega o primeiro valor dos IDs na ordem decrescente (último item adicionado)
                        "SELECT id_equipamento FROM equipamentos ORDER BY id_equipamento DESC LIMIT 1"
                    )
                    id_equip = cur.fetchone()[0]
                    
                cur.execute( # busca o ID do funcionário pelo seu e-mail
                    "SELECT id_funcionario FROM funcionarios WHERE email = %s", (email,)
                )
                id_func = cur.fetchone()[0]
                
                cur.execute(
                    "INSERT INTO operacoes (tipo_operacao, id_funcionario, id_equipamento, quantidade_op) VALUES (%s, %s, %s, %s)", (operacao, id_func, id_equip, quantidade)
                )
                
    
    except Exception as e:
        print("erro 10")
        print(e)
        return

def enviar_email_erro(erro, email_user): # envia e-mail para user detalhando o erro causado no processo
    print("Enviando Email: Erro")
    email = EmailSender(host="smtp.gmail.com", port=587)
    
    email.username = EMAIL
    email.password = EMAIL_PASSWORD
    email.send(
        receivers=[email_user],
        subject="Sem Sucesso",
        text=f"Ocorreu um erro ao processar sua solicitação. Erro {erro}"
    )
    
def enviar_email_sucesso(operacao, equipamento, email_user): #  envia e-mail para user de confirmação
    print("Enviando email: Sucesso")
    email = EmailSender(host="smtp.gmail.com", port=587)
    email.username = EMAIL
    email.password = EMAIL_PASSWORD
    email.send( # email com anexo do arquivo do inventario
        receivers=[EMAIL_ENVIO],
        subject="Sucesso - Inventário",
        attachments={'WinMOD PRO - Inventário.xlsx':Path('WinMOD PRO - Inventário.xlsx')}
    )
    email.send( # email com anexo do arquivo de operacoes
        receivers=[EMAIL_ENVIO],
        subject="Sucesso - Operações",
        attachments={'WinMOD PRO - Operações.xlsx':Path('WinMOD PRO - Operações.xlsx')}
    )
    email.send( # email com anexo do arquivo dos empréstimos
        receivers=[EMAIL_ENVIO],
        subject="Sucesso - Empréstimo",
        attachments={'WinMOD PRO - Empréstimos.xlsx':Path('WinMOD PRO - Empréstimos.xlsx')}
    )
    
    email.send( # email p/ user
        subject=f"Pedido concluído | WinMOD PRO",
        receivers=[email_user],
        text=f"Seu pedido de {operacao} do equipamento {equipamento} foi devidamente concluído."
    )

def gerar_xlsx_operacoes(): # Função para gerar o arquvio excel com as operações feitas no sistema
    try:
        with psycopg.connect(f"dbname=postgres user={DB_USER} password={DB_PASSWORD}") as conn:
            with conn.cursor() as cur:
                cur.execute( # busca de todos IDs presentes na tabela de operações
                    "SELECT id_equipamento FROM operacoes"
                )
                rows_equip = cur.fetchall()
                ids_equip = [row[0] for row in rows_equip] # cria uma tabela temporária pra guardar os atributos respectivos de cada ID
                query = """
                SELECT e.nome, e.id_equipamento 
                FROM UNNEST(%s) AS lista_id 
                JOIN equipamentos e ON e.id_equipamento = lista_id 
                """ # pega dados do nome e fabricante dos itens cujos IDs foram selecionados anteriormente e os coloca em uma tabela temporária "lista_id"
                lista_equip = cur.execute(query, (ids_equip,))
                df_equipamentos = pd.DataFrame(lista_equip)
                
                cur.execute( # busca de todos IDs de funcionarios presentes na tabela de operações
                    "SELECT id_funcionario FROM operacoes"
                )
                rows_funcionarios = cur.fetchall()
                ids_funcionarios = [row[0] for row in rows_funcionarios] # cria uma tabela temporária pra guardar os atributos respectivos de cada ID
                query = """
                SELECT f.nome, f.email 
                FROM UNNEST(%s) AS lista_id 
                JOIN funcionarios f ON f.id_funcionario = lista_id 
                """ # pega dados do nome e e-mail dos funcionários cujos IDs foram selecionados anteriormente e os coloca em uma tabela temporaria "lista_id"
                lista_funcionarios = cur.execute(query, (ids_funcionarios,)) 
                df_funcionarios = pd.DataFrame(lista_funcionarios) # cria dataframe com os dados dos funcionarios
                
                df_operacoes = (pd.read_sql_query("SELECT tipo_operacao, data_operacao, quantidade_op FROM operacoes", con=conn)). reset_index(drop=True) # query: busca dados da tabela de operações e cria um dataframe com esses valores

        # para os dataframes acima, é utilizado o reset_index para evitar erros de sobreposição dos dataframes
               
                conn.commit()
    except Exception as e:
        print(e)

    df_final = pd.concat([df_operacoes, df_equipamentos.reset_index(drop=True), df_funcionarios.reset_index(drop=True)], axis=1) # concatena os dataframes para formar um unico dataframe que será utilizado para escrita no arquivo excel
    
    # CONFIGURAÇÃO DO ARQUIVO EXCEL (TABELAS X COLUNAS)
    wb = Workbook() 
    ws = wb.active
    ws.cell(row=1, column=1, value="Operação")
    ws.cell(row=1, column=2, value="Data")
    ws.cell(row=1, column=3, value="Quantidade")
    ws.cell(row=1, column=4, value="Equipamento")
    ws.cell(row=1, column=5, value="ID")
    ws.cell(row=1, column=6, value="Funcionário")
    ws.cell(row=1, column=7, value="E-mail")
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 7
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 32

    ws.title = "Operações - WinMOD PRO"
    
    for r in dataframe_to_rows(df_final, header=False, index=False):
        ws.append(r)
                
    end_cell = get_column_letter(ws.max_column) + str(ws.max_row) # pega a ultima coluna e linha possível para criar a tabela no tamanho certo

    table_ref = f"A1:{end_cell}"
    
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table = Table(displayName="WinMOD_PRO", ref=table_ref)
    table.tableStyleInfo = style
    
    ws.add_table(table)
    
    file_name = 'WinMOD PRO - Operações.xlsx'
    wb.save(file_name)
        
def gerar_xlsx_emprestimos():  # Função para gerar o arquvio excel com os empréstimos feitas no sistema
    sql_query = """
    SELECT nome, email 
    FROM funcionarios
    INNER JOIN emprestimos
    ON funcionarios.id_funcionario = emprestimos.id_funcionario
    """ # query utilizada para pegar as ocorrências de funcionarios que estão presentes na tabela de emprestimos
    
    with psycopg.connect(f"dbname=postgres user={DB_USER} password={DB_PASSWORD}") as conn:
         
        with conn.cursor() as cur:
                cur.execute(
                    "SELECT id_equipamento FROM operacoes WHERE tipo_operacao = 'Empréstimo'"
                ) # query para resgatar os IDs dos equipamentos que foram emprestados
                rows = cur.fetchall()
                ids = [row[0] for row in rows] # cria uma tabela temporária pra guardar os atributos respectivos de cada ID
                query = """
                SELECT e.nome, e.categoria 
                FROM UNNEST(%s) AS lista_id 
                JOIN equipamentos e ON e.id_equipamento = lista_id 
                """ # cria uma lista temporaria pra guardar os atributos respectivos de cada ID
                cur.execute(query, (ids, ))
                lista = cur.fetchall()
                conn.commit()
                    
        df_operacoes = pd.DataFrame(lista) # cria um dataframe com os dados das operações
        query = """
        SELECT nome, categoria 
        FROM equipamentos
        WHERE id_equipamento IN (
            SELECT id_equipamento FROM operacoes WHERE tipo_operacao = 'Empréstimo'
        ) """ # busca o nome e a categoria dos equipamentos que constam como "Empréstimo" em seu campo de tipo de operação

        df_emprestimos = (pd.read_sql_query("SELECT data_retorno, quantidade, status_atual FROM emprestimos", con=conn)).reset_index(drop=True) # query: busca dados da tabela de empréstimos e cria um dataframe com esses valores
        df_funcionarios = (pd.read_sql_query(sql_query, con=conn)).reset_index(drop=True)
        df_operacoes = df_operacoes.reset_index(drop=True) 

        # para os dataframes acima, é utilizado o reset_index para evitar erros de sobreposição dos dataframes
        
        df_final = pd.concat([df_operacoes, df_funcionarios, df_emprestimos], axis=1) # concatena os dataframes para formar um unico dataframe que será utilizado para construir o arquivo excel
        
        # CONFIGURAÇÃO DO ARQUIVO EXCEL (TABELAS X COLUNAS)
        wb = Workbook() 
        ws = wb.active
        ws.cell(row=1, column=1, value="Equipamento")
        ws.cell(row=1, column=2, value="Categoria")
        ws.cell(row=1, column=3, value="Funcionário")
        ws.cell(row=1, column=4, value="E-mail")
        ws.cell(row=1, column=5, value="Data p/ retorno")
        ws.cell(row=1, column=6, value="Quantidade")
        ws.cell(row=1, column=7, value="Status Atual")
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 32
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 17
        ws.column_dimensions['F'].width = 13
        ws.column_dimensions['G'].width = 14

        ws.title = "Empréstimos - WinMOD PRO"
        
        for r in dataframe_to_rows(df_final, header=False, index=False):
            ws.append(r)
                    
        end_cell = get_column_letter(ws.max_column) + str(ws.max_row) # pega a ultima coluna e linha possível para criar a tabela no tamanho certo
    
        table_ref = f"A1:{end_cell}"
        
        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        table = Table(displayName="WinMOD_PRO", ref=table_ref)
        table.tableStyleInfo = style
        
        ws.add_table(table)
        
        file_name = 'WinMOD PRO - Empréstimos.xlsx'
        wb.save(file_name) 

def gerar_xlsx():
    
    sql_query = "SELECT * FROM equipamentos WHERE ativo = true" # seleciona todos os itens ativos na tabela de equipamentos
    
    with psycopg.connect(f"dbname=postgres user={DB_USER} password={DB_PASSWORD}") as conn:
        df = pd.read_sql_query(sql_query, con=conn) # criação do dataframe utilizando a query acima

        # CONFIGURAÇÃO DO ARQUIVO EXCEL (TABELAS X COLUNAS)
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="ID")
        ws.cell(row=1, column=2, value="Nome")
        ws.cell(row=1, column=3, value="Quantidade")
        ws.cell(row=1, column=4, value="Descrição")
        ws.cell(row=1, column=5, value="Fabricante")
        ws.cell(row=1, column=6, value="Nº Série/Nº Modelo")
        ws.cell(row=1, column=7, value="Localização")
        ws.cell(row=1, column=8, value="Status")
        ws.cell(row=1, column=9, value="Categoria")
        ws.cell(row=1, column=10, value="Ativo")
        ws.column_dimensions['B'].width = 56
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 120
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 28
        ws.column_dimensions['G'].width = 13
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 34
        ws.column_dimensions['J'].width = 13
        ws.title = "Inventário WinMOD PRO"
        
        for row in df.values.tolist():
            ws.append(row)
            
        end_cell = get_column_letter(ws.max_column) + str(ws.max_row)
        table_ref = f"A1:{end_cell}"
        
        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        table = Table(displayName="WinMOD_PRO", ref=table_ref)
        table.tableStyleInfo = style
        
        ws.add_table(table)
        
        file_name = 'WinMOD PRO - Inventário.xlsx'
        wb.save(file_name) 

while True:
    if esperar_email():
        print("Processando anexo")
        ler_anexo()
        
    time.sleep(1)
    
    

"""
TO DO:

tratamento do nome da função nome_ja_cadastrado, a fim de evitar itens duplicados no bd por causa de uma letra maiúscula
Corrigir formatação do timestamp para melhor visualização

"""